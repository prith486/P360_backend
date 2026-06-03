import uuid
from typing import Optional, Any, Dict
from decimal import Decimal
from io import BytesIO, StringIO
import csv

from fastapi import APIRouter, Depends, HTTPException, status, BackgroundTasks, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy.orm import joinedload
from openpyxl import Workbook
from openpyxl.styles import Font

from app.core.database import get_db
from app.api.deps import get_current_faculty, get_current_student, get_current_active_user
from app.models.faculty import Faculty
from app.models.student import Student
from app.models.auth_user import AuthUser
from app.models.assessment import Assessment
from app.models.assessment_question import AssessmentQuestion
from app.models.assessment_attempt import AssessmentAttempt
from app.models.question import Question
from app.schemas.assessment import (
    AssessmentCreate, AssessmentUpdate, AssessmentRead, AssessmentAssigned,
    AssessmentQuestionAdd
)
from app.schemas.question import QuestionRead
from app.models.submission import Submission
from app.models.enums import SubmissionStatus, QuestionType, AssessmentType, ProgrammingLanguage, DifficultyLevel
from app.schemas.assessment_attempt import (
    AttemptStartRequest, AttemptSubmitRequest,
    AssessmentAttemptRead, AssessmentAttemptDetailed
)
from app.schemas.ai_generation import (
    GenerateTestRequest,
    GenerateAssessmentConfirmRequest,
)
from app.schemas.submission import SubmissionRead
from datetime import datetime, timezone, timedelta
from app.core.cache import get_cached, set_cached, invalidate_cache
from app.integrations.judge0 import judge0_client
from app.services.ai_service import generate_test as ai_generate_test

router = APIRouter()

from app.core.logging_config import get_logger
logger = get_logger(__name__)

def normalize_branch(branch: str) -> str:
    """Normalize branch names for consistent eligibility checks."""
    if not branch:
        return ""
    mapping = {
        "CS": "computer_science",
        "IT": "information_technology", 
        "ECE": "electronics",
        "Mech": "electrical",
        "Civil": "civil",
        "computer_science": "computer_science",
        "information_technology": "information_technology",
    }
    return mapping.get(branch, branch.lower().replace(" ", "_"))

def normalize_year(year: str) -> str:
    """Normalize batch years to academic year enums."""
    if not year:
        return ""
    # Map batch years to academic year enums
    # Assuming current year is 2024-25
    batch_to_year = {
        "2021": "fourth",
        "2022": "third",
        "2023": "second", 
        "2024": "first",
        "first": "first",
        "second": "second",
        "third": "third",
        "fourth": "fourth",
    }
    return batch_to_year.get(str(year).lower(), year.lower())

def format_assessment_dict(a: Assessment, db=None, current_user_id=None) -> dict:
    """
    Standardized formatter to ensure status logic is flawless across all views.
    Handles auto-promotion from 'draft' to 'scheduled' or 'active' based on current time.
    When db is provided, resolves creator name from Faculty + AuthUser tables.
    """
    from datetime import timedelta
    
    # 1. Base status normalization
    status_val = a.status.lower() if a.status else "draft"
    
    # 2. Visually "UPGRADE" status based on time and flags
    now = datetime.utcnow()
    start = a.start_time
    
    if start and start.tzinfo is not None:
        # If start_time is timezone aware, make now aware too
        now = datetime.now(timezone.utc)
    elif start and now.tzinfo is not None:
        # If now is aware but start is naive (shouldn't happen with utcnow), make start aware or now naive
        now = now.replace(tzinfo=None)

    if a.is_published:
        if status_val == "draft":
            status_val = "scheduled"
            
        if start:
            if now >= start:
                # Calculate end_time
                duration = a.duration_minutes or 0
                end = a.end_time
                if not end:
                    end = start + timedelta(minutes=duration)
                
                if now < end:
                    status_val = "active"
                else:
                    status_val = "completed"
            else:
                status_val = "scheduled"
    
    # 3. Final mapping to frontend-friendly strings
    if status_val == "published":
        status_val = "scheduled"
    
    # 4. Resolve creator info (only when db is passed - avoids N+1 in lists by pre-fetching)
    creator_name = None
    creator_employee_id = None
    creator_department = None
    if db is not None and a.created_by:
        try:
            fac = db.query(Faculty).filter(Faculty.user_id == a.created_by).first()
            if fac:
                auth_user = db.query(AuthUser).filter(AuthUser.id == a.created_by).first()
                creator_name = auth_user.full_name if auth_user else None
                creator_employee_id = fac.employee_id
                creator_department = fac.department
        except Exception:
            pass
        
    res = {
        "id": str(a.id),
        "title": a.title.strip() if a.title else "Untitled Assessment",
        "description": a.description,
        "status": status_val,
        "is_published": a.is_published,
        "created_by": str(a.created_by),
        "created_by_name": creator_name,
        "created_by_employee_id": creator_employee_id,
        "created_by_department": creator_department,
        "created_at": a.created_at.isoformat() if a.created_at else None,
        "type": a.assessment_type.value if hasattr(a.assessment_type, 'value') else str(a.assessment_type),
        "assessment_type": a.assessment_type.value if hasattr(a.assessment_type, 'value') else str(a.assessment_type),
        "duration": a.duration_minutes,
        "duration_minutes": a.duration_minutes,
        "maxScore": a.total_marks,
        "max_score": a.total_marks,
        "total_marks": a.total_marks,
        "passingScore": a.passing_marks,
        "passing_score": a.passing_marks,
        "passing_marks": a.passing_marks,
        "departments": a.target_branches,
        "target_branches": a.target_branches,
        "batches": a.target_years,
        "target_years": a.target_years,
        "shuffleQuestions": a.shuffle_questions,
        "shuffle_questions": a.shuffle_questions,
        "showResultsImmediately": a.show_results_immediately,
        "show_results_immediately": a.show_results_immediately,
        "minCGPA": float(a.min_cgpa) if a.min_cgpa else None,
        "min_cgpa": float(a.min_cgpa) if a.min_cgpa else None,
        "totalQuestions": a.total_questions,
        "total_questions": a.total_questions,
        "track_tab_switches": a.track_tab_switches,
        "trackTabSwitches": a.track_tab_switches,
        "tab_switch_limit": a.tab_switch_limit,
        "tabSwitchLimit": a.tab_switch_limit,
        "slug": a.slug,
        "is_active": a.is_active,
        "date": a.start_time.strftime("%Y-%m-%d") if a.start_time else None,
        "time": a.start_time.strftime("%H:%M:%S") if a.start_time else None,
        "start_time": a.start_time.isoformat() if a.start_time else None,
        "end_time": a.end_time.isoformat() if a.end_time else None,
        "registrationDeadline": a.registration_deadline.strftime("%Y-%m-%dT%H:%M:%SZ") if a.registration_deadline else None,
        "registration_deadline": a.registration_deadline.isoformat() if a.registration_deadline else None,
        "max_attempts": a.max_attempts,
        "instructions": a.instructions,
        "enable_proctoring": a.enable_proctoring,
        "tab_switch_limit": a.tab_switch_limit if hasattr(a, 'tab_switch_limit') else 3,
    }
    # Inject is_mine only when we know the current faculty
    if current_user_id is not None:
        res["is_mine"] = str(a.created_by) == str(current_user_id)
    return res


def _slug_from_text(seed_text: str) -> str:
    slug_uuid = str(uuid.uuid4())[:8]
    cleaned = "-".join(seed_text.lower().strip().split())
    cleaned = cleaned[:80] if cleaned else "question"
    return f"{cleaned}-{slug_uuid}"


def _create_question_from_generated_payload(data: dict[str, Any], faculty_user_id: uuid.UUID) -> Question:
    question_text = str(data.get("question_text", "")).strip()
    if not question_text:
        raise ValueError("question_text is required")

    q_type = str(data.get("type", "")).strip().lower()
    if q_type not in {"mcq", "coding"}:
        raise ValueError("type must be mcq or coding")

    difficulty_val = str(data.get("difficulty", "medium")).strip().lower()
    if difficulty_val not in {"easy", "medium", "hard", "expert"}:
        raise ValueError("difficulty must be easy, medium, hard, or expert")

    payload: dict[str, Any] = {
        "title": question_text[:120] if len(question_text) > 120 else question_text,
        "slug": _slug_from_text(question_text),
        "difficulty": DifficultyLevel(difficulty_val),
        "question_type": QuestionType(q_type),
        "description": question_text,
        "tags": data.get("tags") if isinstance(data.get("tags"), list) else [],
        "editorial": data.get("explanation"),
        "created_by": faculty_user_id,
        "total_submissions": 0,
        "total_accepted": 0,
        "acceptance_rate": 0.0,
        "is_active": True,
        "source": "AI Generated",
    }

    if q_type == "mcq":
        options = data.get("options") if isinstance(data.get("options"), dict) else {}
        option_map = {
            "A": str(options.get("A", "")).strip(),
            "B": str(options.get("B", "")).strip(),
            "C": str(options.get("C", "")).strip(),
            "D": str(options.get("D", "")).strip(),
        }
        if not all(option_map.values()):
            raise ValueError("mcq requires non-empty options A/B/C/D")

        correct_option = str(data.get("correct_option", "")).strip().upper()
        if correct_option not in {"A", "B", "C", "D"}:
            raise ValueError("mcq requires correct_option in A/B/C/D")

        labels = ["A", "B", "C", "D"]
        formatted_options = []
        correct_index = None
        for idx, label in enumerate(labels, start=1):
            formatted_options.append(
                {
                    "id": idx,
                    "label": label,
                    "text": option_map[label],
                    "is_correct": label == correct_option,
                }
            )
            if label == correct_option:
                correct_index = idx

        payload["options"] = formatted_options
        payload["solution_code"] = {"correct_option": correct_index}

    if q_type == "coding":
        payload["starter_code"] = data.get("starter_code") if isinstance(data.get("starter_code"), dict) else {}

        sample_cases = []
        if isinstance(data.get("sample_test_cases"), list):
            for case in data.get("sample_test_cases", []):
                if isinstance(case, dict):
                    sample_cases.append(
                        {
                            "input": str(case.get("input", "")).strip(),
                            "output": str(case.get("expected_output", case.get("output", ""))).strip(),
                        }
                    )
        payload["sample_test_cases"] = sample_cases

    return Question(**payload)


def _build_result_row(rank: int, attempt: AssessmentAttempt) -> dict[str, Any]:
    student = attempt.student
    auth_user = student.user if student else None
    total_score = float(attempt.total_score) if attempt.total_score is not None else 0.0
    max_score = float(attempt.max_score) if attempt.max_score is not None else 0.0
    percentage = float(attempt.percentage) if attempt.percentage is not None else (round((total_score / max_score) * 100, 2) if max_score > 0 else 0.0)

    return {
        "Rank": rank,
        "Student Name": auth_user.full_name if auth_user else "Unknown",
        "Roll No": student.roll_number if student else "N/A",
        "Branch": student.branch.value if student and hasattr(student.branch, "value") else (str(student.branch) if student and student.branch else "N/A"),
        "Year": student.current_year.value if student and hasattr(student.current_year, "value") else (str(student.current_year) if student and student.current_year else "N/A"),
        "CGPA": float(student.cgpa) if student and student.cgpa is not None else None,
        "Total Score": total_score,
        "Max Score": max_score,
        "Percentage": percentage,
        "Pass/Fail": "Pass" if attempt.is_passed else "Fail",
        "Time Taken (minutes)": attempt.time_taken_minutes,
        "Tab Switches": attempt.tab_switch_count,
        "Submitted At": attempt.submitted_at.isoformat() if attempt.submitted_at else None,
    }


@router.post("", response_model=dict)
@router.post("/", response_model=dict, include_in_schema=False)
def create_assessment(
    data: AssessmentCreate,
    db: Session = Depends(get_db),
    current_faculty: Faculty = Depends(get_current_faculty)
):
    cleaned_title = data.title.strip()
    slug_uuid = "{:.8}".format(str(uuid.uuid4()))
    slug = f"{cleaned_title.lower().replace(' ', '-')}-{slug_uuid}"
    
    # Try to map to the new DB fields and calculate the date
    from app.models.enums import AssessmentType
    try:
        a_type = AssessmentType(data.type)
    except ValueError:
        a_type = AssessmentType.MIXED # fallback
        
    # Parse start_time: prefer ISO string start_time/end_time fields first,
    # then fall back to legacy date+time string pair
    from dateutil.parser import parse as parse_dt
    start_dt = None
    end_dt = None
    
    if data.start_time:
        try:
            start_dt = parse_dt(data.start_time)
        except Exception:
            pass
    elif data.date and data.time:
        try:
            start_dt = parse_dt(f"{data.date} {data.time}")
        except Exception:
            pass
            
    if data.end_time:
        try:
            end_dt = parse_dt(data.end_time)
        except Exception:
            pass
            
    reg_dl = None
    if data.registrationDeadline:
        try:
            reg_dl = parse_dt(data.registrationDeadline)
        except Exception:
            pass
            
    # Rule: Always create as draft regardless of what start_time or is_published is provided
    is_published = False
    target_status = "draft"
    
    assessment = Assessment(
        title=cleaned_title,
        description=data.description,
        assessment_type=a_type,
        duration_minutes=data.duration,
        start_time=start_dt,
        end_time=end_dt,
        registration_deadline=reg_dl,
        total_marks=data.maxScore,
        passing_marks=data.passingScore,
        target_branches=data.departments,
        target_years=data.batches,
        shuffle_questions=data.shuffleQuestions,
        show_results_immediately=data.showResultsImmediately,
        min_cgpa=data.minCGPA,
        status=target_status,
        slug=slug,
        created_by=current_faculty.user_id,
        total_questions=0,
        total_attempts=0,
        is_active=True,
        is_published=is_published,
        enable_proctoring=data.enable_proctoring,
        track_tab_switches=data.track_tab_switches,
        instructions=data.instructions,
        max_attempts=data.max_attempts
    )
    db.add(assessment)
    db.commit()
    db.refresh(assessment)
    
    # Aggressive cache invalidation:
    # 1. All assessment lists (faculty & general)
    invalidate_cache("assessments_")
    # 2. All student dashboards (since new assessments might appear there)
    invalidate_cache("student_dashboard:")
    # 3. Faculty dashboards
    invalidate_cache("faculty_dashboard:")
    
    # Use centralized formatter for flawless response
    return {
        "success": True,
        "data": format_assessment_dict(assessment)
    }

@router.put("/{assessment_id}", response_model=dict)
def update_assessment(
    assessment_id: uuid.UUID,
    data: AssessmentUpdate,
    db: Session = Depends(get_db),
    current_faculty: Faculty = Depends(get_current_faculty)
):
    """Update an existing assessment (Faculty only)."""
    assessment = db.query(Assessment).filter(Assessment.id == assessment_id).first()
    if not assessment:
        raise HTTPException(status_code=404, detail="Assessment not found")
        
    if str(assessment.created_by) != str(current_faculty.user_id):
        raise HTTPException(status_code=403, detail="Not authorized to edit this assessment")

    # Update basic string fields if provided
    if data.title is not None:
        assessment.title = data.title.strip()
    if data.description is not None:
        assessment.description = data.description
        
    try:
        if data.type is not None:
            a_type_val = data.type.value if hasattr(data.type, 'value') else str(data.type)
            assessment.assessment_type = AssessmentType(a_type_val)
    except ValueError:
        pass
        
    if data.duration is not None:
        assessment.duration_minutes = data.duration
        
    from dateutil.parser import parse as parse_dt
    
    if data.start_time:
        try:
            assessment.start_time = parse_dt(data.start_time)
        except Exception:
            pass
    elif data.date and data.time:
        try:
            assessment.start_time = parse_dt(f"{data.date} {data.time}")
        except Exception:
            pass
            
    if data.end_time:
        try:
            assessment.end_time = parse_dt(data.end_time)
        except Exception:
            pass
            
    if data.registrationDeadline:
        try:
            assessment.registration_deadline = parse_dt(data.registrationDeadline)
        except Exception:
            pass
            
    if data.maxScore is not None:
        assessment.total_marks = data.maxScore
    if data.passingScore is not None:
        assessment.passing_marks = data.passingScore
    if data.departments is not None:
        assessment.target_branches = data.departments
    if data.batches is not None:
        assessment.target_years = data.batches
    if data.shuffleQuestions is not None:
        assessment.shuffle_questions = data.shuffleQuestions
    if data.showResultsImmediately is not None:
        assessment.show_results_immediately = data.showResultsImmediately
    if data.minCGPA is not None:
        assessment.min_cgpa = data.minCGPA
    if data.enable_proctoring is not None:
        assessment.enable_proctoring = data.enable_proctoring
    if data.track_tab_switches is not None:
        assessment.track_tab_switches = data.track_tab_switches
    if data.instructions is not None:
        assessment.instructions = data.instructions
    if data.max_attempts is not None:
        assessment.max_attempts = data.max_attempts
    if data.is_published is not None:
        assessment.is_published = data.is_published

    # Update state based on status and time
    now = datetime.utcnow()
    
    start_time_cmp = assessment.start_time
    if start_time_cmp and start_time_cmp.tzinfo is not None:
        start_time_cmp = start_time_cmp.replace(tzinfo=None)
        
    end_time_cmp = assessment.end_time
    if end_time_cmp and end_time_cmp.tzinfo is not None:
        end_time_cmp = end_time_cmp.replace(tzinfo=None)
        
    if not assessment.is_published:
        assessment.status = "draft"
    elif start_time_cmp and start_time_cmp > now:
        assessment.status = "scheduled"
    elif (start_time_cmp is None or start_time_cmp <= now) and (end_time_cmp is None or end_time_cmp >= now):
        assessment.status = "active"
    else:
        assessment.status = "completed"

    db.commit()
    db.refresh(assessment)
    
    # Aggressive cache invalidation
    invalidate_cache(f"assessment_details_{str(assessment.id)}")
    invalidate_cache("assessments_")
    invalidate_cache("student_dashboard:")
    invalidate_cache("faculty_dashboard:")
    
    return {
        "success": True,
        "data": format_assessment_dict(assessment)
    }

@router.get("", response_model=dict)
@router.get("/", response_model=dict, include_in_schema=False)
def get_faculty_assessments(
    status_filter: Optional[str] = None,
    page: Optional[int] = Query(None, description="Page number for pagination"),
    limit: Optional[int] = Query(10, description="Items per page"),
    search: Optional[str] = Query(None, description="Search query"),
    status_query: Optional[str] = Query(None, alias="status", description="Filter by status (draft, scheduled, active, completed)"),
    type: Optional[str] = Query(None, description="Filter by type"),
    mine_only: Optional[bool] = Query(None, description="If true, only return assessments created by the logged-in faculty"),
    db: Session = Depends(get_db),
    current_faculty: Faculty = Depends(get_current_faculty)
):
    """
    Returns ALL assessments visible to faculty.
    Each item has:
      - is_mine: True if created by the logged-in faculty, False otherwise
      - created_by_name: Full name of the creator faculty
      - created_by_employee_id: Employee ID of creator
      - created_by_department: Department of creator
    Frontend can use is_mine to show ownership, filter views, or restrict edit/delete actions.
    """
    # Disable cache for this endpoint since it now shows all assessments with per-user flags
    # (cache would mix results for different faculty)

    from sqlalchemy.orm import selectinload, joinedload
    from sqlalchemy import or_

    # --- Fetch ALL assessments, not just the current faculty's ---
    query = db.query(Assessment).options(selectinload(Assessment.attempts))
    
    # Optional mine_only filter
    if mine_only:
        query = query.filter(Assessment.created_by == current_faculty.user_id)
    
    # Backward compatibility for status_filter
    if status_filter == "published":
        query = query.filter(Assessment.is_published == True)
    elif status_filter == "draft":
        # Drafts are only visible to the creator
        query = query.filter(
            Assessment.is_published == False,
            Assessment.created_by == current_faculty.user_id
        )
        
    if search:
        query = query.filter(or_(Assessment.title.ilike(f"%{search}%"), Assessment.description.ilike(f"%{search}%")))
    if type:
        query = query.filter(Assessment.assessment_type == type)
        
    query = query.order_by(Assessment.created_at.desc())
    assessments = query.all()
    
    # --- Batch-load creator info to avoid N+1 queries ---
    # Collect unique creator UUIDs
    creator_ids = list({a.created_by for a in assessments if a.created_by})
    
    # Bulk fetch faculty + auth user rows
    faculty_rows = db.query(Faculty, AuthUser).join(
        AuthUser, Faculty.user_id == AuthUser.id
    ).filter(Faculty.user_id.in_(creator_ids)).all()
    
    # Build lookup map: user_id -> {name, employee_id, department}
    creator_map = {}
    for fac, auth in faculty_rows:
        creator_map[str(fac.user_id)] = {
            "name": auth.full_name,
            "employee_id": fac.employee_id,
            "department": fac.department
        }
    
    current_fac_uid = str(current_faculty.user_id)
    
    data = []
    for a in assessments:
        item = format_assessment_dict(a, current_user_id=current_fac_uid)
        
        # Hide other faculty's drafts
        if not a.is_published and str(a.created_by) != current_fac_uid:
            continue
        
        if status_query and item["status"] != status_query.lower():
            continue
        
        # Inject creator info from our pre-loaded map (avoids N+1)
        creator_info = creator_map.get(str(a.created_by), {})
        item["created_by_name"] = creator_info.get("name")
        item["created_by_employee_id"] = creator_info.get("employee_id")
        item["created_by_department"] = creator_info.get("department")
        
        attempts = a.attempts
        item["participants"] = len(set([str(att.student_id) for att in attempts]))
        
        scores = [float(att.total_score) for att in attempts if att.is_submitted and att.total_score is not None]
        item["avgScore"] = round(float(sum(scores) / len(scores)), 2) if scores else 0.0
            
        data.append(item)
        
    if page is not None:
        total = len(data)
        limit = limit or 10
        start = (int(page) - 1) * int(limit)
        end = start + int(limit)
        paginated_data = data[start:end]
        return {
            "success": True,
            "data": paginated_data,
            "meta": {
                "total": total,
                "page": page,
                "limit": limit,
                "pages": (total + limit - 1) // limit if limit else 1
            }
        }
    
    return {"success": True, "data": data}


@router.get("/my", response_model=dict)
def get_my_assessments(
    page: Optional[int] = Query(None, description="Page number for pagination"),
    limit: Optional[int] = Query(10, description="Items per page"),
    search: Optional[str] = Query(None, description="Search query"),
    status_query: Optional[str] = Query(None, alias="status", description="Filter by status (scheduled, active, completed)"),
    type: Optional[str] = Query(None, description="Filter by type"),
    db: Session = Depends(get_db),
    current_student: Student = Depends(get_current_student)
):
    branch_val = current_student.branch.value if hasattr(current_student.branch, 'value') else str(current_student.branch)
    year_val = current_student.current_year.value if hasattr(current_student.current_year, 'value') else str(current_student.current_year)
    
    cache_key = None
    if not page and not search and not status_query and not type:
        cache_key = f"assessments_stu_{str(current_student.user_id)}_{branch_val}_{year_val}"
        cached = get_cached(cache_key)
        if cached:
            return cached

    from sqlalchemy import or_
    query = db.query(Assessment).filter(
        Assessment.is_published == True,
        Assessment.is_active == True
    )
    
    if search:
        query = query.filter(or_(Assessment.title.ilike(f"%{search}%"), Assessment.description.ilike(f"%{search}%")))
    if type:
        query = query.filter(Assessment.assessment_type == type)
        
    query = query.order_by(Assessment.start_time.asc())
    assessments = query.all()
    
    # Normalize branch and year for comparison with frontend labels
    branch_map = {
        "computer_science": ["CS", "Computer Science", "CSE"],
        "information_technology": ["IT", "Information Technology"],
        "electronics": ["ECE", "Electronics", "ENTC"],
        "mechanical": ["Mech", "Mechanical"],
        "civil": ["Civil"],
        "electrical": ["EE", "Electrical"],
        "ai_ml": ["AI/ML", "AI", "ML"]
    }
    
    year_map = {
        "first": ["First Year", "1st Year"],
        "second": ["Second Year", "2nd Year"],
        "third": ["Third Year", "3rd Year"],
        "fourth": ["Fourth Year", "Final Year", "4th Year"]
    }
    
    student_branches = [branch_val]
    if branch_val in branch_map:
        student_branches.extend(branch_map[branch_val])
        
    student_years = [year_val, str(current_student.batch_year)]
    if year_val in year_map:
        student_years.extend(year_map[year_val])
    
    # 1. First pass: Filter assessments the student is eligible for
    from app.utils.helpers import normalize_branch, normalize_year
    eligible_assessments = []
    for a in assessments:
        # Check if any of student's possible branch labels match target branches
        student_branch = normalize_branch(branch_val)
        assessment_branches = [normalize_branch(b) for b in (a.target_branches or [])]
        
        b_match = not assessment_branches
        if assessment_branches:
            if student_branch in assessment_branches:
                b_match = True
        
        # Also check other labels for safety
        if not b_match:
            for b in student_branches:
                if normalize_branch(b) in assessment_branches:
                    b_match = True
                    break
        
        # Check if any of student's possible year/batch labels match target years
        student_year = normalize_year(year_val)
        assessment_years = [normalize_year(y) for y in (a.target_years or [])]
        
        y_match = not assessment_years
        if assessment_years:
            if student_year in assessment_years:
                y_match = True
        
        # Fallback check original student labels
        if not y_match:
            for y in student_years:
                if normalize_year(y) in assessment_years:
                    y_match = True
                    break
            
        if b_match and y_match:
            eligible_assessments.append(a)

    if not eligible_assessments:
        res = {"success": True, "data": []}
        if cache_key:
            set_cached(cache_key, res, 60)
        return res

    # 2. Batch fetch ALL attempts for all eligible assessments in ONE query
    eligible_ids = [a.id for a in eligible_assessments]
    all_attempts_list = db.query(AssessmentAttempt).filter(
        AssessmentAttempt.assessment_id.in_(eligible_ids)
    ).all()

    # --- Batch-load creator info to avoid N+1 queries ---
    creator_ids = list({a.created_by for a in eligible_assessments if a.created_by})
    from app.models.faculty import Faculty
    from app.models.auth_user import AuthUser
    faculty_rows = db.query(Faculty, AuthUser).join(
        AuthUser, Faculty.user_id == AuthUser.id
    ).filter(Faculty.user_id.in_(creator_ids)).all()
    
    creator_map = {}
    for fac, auth in faculty_rows:
        creator_map[str(fac.user_id)] = auth.full_name

    # 3. Group attempts by assessment_id for in-memory processing
    from collections import defaultdict
    attempts_by_assessment = defaultdict(list)
    for att in all_attempts_list:
        attempts_by_assessment[att.assessment_id].append(att)

    # 4. Final pass: Format with in-memory calculations
    valid_assessments = []
    for a in eligible_assessments:
        attempts = attempts_by_assessment[a.id]
        
        # Find current student's submitted attempt
        attempt = next((
            att for att in attempts 
            if att.student_id == current_student.id and att.is_submitted
        ), None)
            
        total_participants = len(set(att.student_id for att in attempts))
            
        # calculate score result
        res_str = None
        if attempt:
            if a.passing_marks is not None and attempt.total_score is not None:
                res_str = "pass" if attempt.total_score >= a.passing_marks else "fail"
            else:
                res_str = "pass" if attempt.total_score is not None and attempt.total_score > 0 else "fail"
                    
        # Calculate rank dynamically from preloaded attempts
        my_rank = None
        if attempt:
            # Count how many students have a strictly higher score (in-memory)
            higher_scores = sum(
                1 for att in attempts 
                if att.is_submitted and att.total_score is not None and att.total_score > attempt.total_score
            )
            my_rank = higher_scores + 1

        # Use centralized formatter
        formatted = format_assessment_dict(a)
        formatted["created_by_name"] = creator_map.get(str(a.created_by))
        
        if status_query and formatted["status"] != status_query.lower():
            continue
            
        # Map student-specific fields
        assigned = {
            **formatted,
            "totalParticipants": total_participants,
            "attempted": True if attempt else False,
            "attempt_id": str(attempt.id) if attempt else None,
            "score": float(attempt.total_score) if attempt and attempt.total_score is not None else None,
            "result": res_str,
            "timeTaken": int((attempt.submitted_at - attempt.started_at).total_seconds() / 60) if attempt and attempt.submitted_at and attempt.started_at else None,
            "rank": my_rank
        }
        valid_assessments.append(assigned)
        
    if page is not None:
        total = len(valid_assessments)
        limit = limit or 10
        start = (page - 1) * limit
        end = start + limit
        paginated_data = valid_assessments[start:end]
        return {
            "success": True,
            "data": paginated_data,
            "meta": {
                "total": total,
                "page": page,
                "limit": limit,
                "pages": (total + limit - 1) // limit if limit else 1
            }
        }
            
    res = {"success": True, "data": valid_assessments}
    if cache_key:
        set_cached(cache_key, res, 60)
    return res


@router.post("/generate", response_model=dict)
def generate_assessment_proposal(
    payload: GenerateTestRequest,
    db: Session = Depends(get_db),
    current_faculty: Faculty = Depends(get_current_faculty),
):
    _ = current_faculty
    proposal = ai_generate_test(payload.model_dump(), db)
    return {"success": True, "data": proposal}


@router.post("/generate/confirm", response_model=dict)
def confirm_generated_assessment(
    payload: GenerateAssessmentConfirmRequest,
    db: Session = Depends(get_db),
    current_faculty: Faculty = Depends(get_current_faculty),
):
    meta = payload.assessment_meta

    from dateutil.parser import parse as parse_dt

    start_dt = None
    end_dt = None
    if meta.start_time:
        try:
            start_dt = parse_dt(meta.start_time)
        except Exception:
            start_dt = None
    if meta.end_time:
        try:
            end_dt = parse_dt(meta.end_time)
        except Exception:
            end_dt = None

    saved_new_questions: list[Question] = []
    for item in payload.new_questions_to_save:
        try:
            question = _create_question_from_generated_payload(item, current_faculty.user_id)
            db.add(question)
            db.flush()
            saved_new_questions.append(question)
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Invalid generated question payload: {exc}")

    selected_ids: list[uuid.UUID] = []

    for raw_id in payload.selected_bank_question_ids:
        try:
            selected_ids.append(uuid.UUID(str(raw_id)))
        except Exception:
            continue

    saved_new_id_map = {str(q.id): q for q in saved_new_questions}
    ordered_ids_raw = payload.question_order or (payload.selected_bank_question_ids + list(saved_new_id_map.keys()))
    ordered_ids: list[uuid.UUID] = []
    for raw_id in ordered_ids_raw:
        try:
            parsed_id = uuid.UUID(str(raw_id))
        except Exception:
            continue
        ordered_ids.append(parsed_id)

    if not ordered_ids:
        ordered_ids = selected_ids + [q.id for q in saved_new_questions]

    questions = db.query(Question).filter(Question.id.in_(ordered_ids)).all()
    question_map = {str(q.id): q for q in questions}

    cleaned_title = meta.title.strip()
    assessment = Assessment(
        title=cleaned_title,
        description=meta.description or "AI Generated Assessment",
        assessment_type=AssessmentType.MIXED,
        duration_minutes=meta.duration,
        start_time=start_dt,
        end_time=end_dt,
        total_marks=0,
        passing_marks=meta.passing_score or 0,
        target_branches=meta.branch,
        target_years=meta.year,
        shuffle_questions=False,
        show_results_immediately=False,
        status="draft",
        slug=f"{cleaned_title.lower().replace(' ', '-')}-{str(uuid.uuid4())[:8]}",
        created_by=current_faculty.user_id,
        total_questions=0,
        total_attempts=0,
        is_active=True,
        is_published=False,
        enable_proctoring=bool((meta.proctoring_config or {}).get("enabled", False)),
        track_tab_switches=bool(
            (meta.proctoring_config or {}).get(
                "track_tab_switches",
                (meta.proctoring_config or {}).get("trackTabSwitches", False),
            )
        ),
        tab_switch_limit=int(
            (meta.proctoring_config or {}).get(
                "tab_switch_limit",
                (meta.proctoring_config or {}).get("tabSwitchLimit", 3),
            )
            or 3
        ),
        instructions=(meta.proctoring_config or {}).get("instructions") if meta.proctoring_config else None,
    )
    db.add(assessment)
    db.flush()

    total_marks = 0
    created_mappings = 0
    for order, qid in enumerate(ordered_ids, start=1):
        question = question_map.get(str(qid))
        if not question:
            continue

        marks = int(question.max_score or 100)
        mapping = AssessmentQuestion(
            assessment_id=assessment.id,
            question_id=question.id,
            question_order=order,
            marks=marks,
        )
        db.add(mapping)
        total_marks += marks
        created_mappings += 1

    assessment.total_questions = created_mappings
    assessment.total_marks = total_marks

    db.commit()
    db.refresh(assessment)

    invalidate_cache("assessment_")
    invalidate_cache("assessments_")
    invalidate_cache("student_dashboard:")
    invalidate_cache("faculty_dashboard:")
    invalidate_cache("questions_")

    return {
        "success": True,
        "assessment_id": str(assessment.id),
        "status": "draft",
    }

@router.get("/{assessment_id}", response_model=dict)
def get_assessment(
    assessment_id: uuid.UUID,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_active_user)
):
    cache_key = f"assessment_details_{str(assessment_id)}_{str(current_user['user_id'])}"
    cached = get_cached(cache_key)
    if cached:
        return cached
 
    assessment = db.query(Assessment).filter(Assessment.id == assessment_id).first()
    if not assessment:
        raise HTTPException(status_code=404, detail="Assessment not found")
        
    role = current_user["role"]
    user_id_str = str(current_user["user_id"])
    
    if user_id_str == "00000000-0000-0000-0000-000000000002":
        role = "faculty"
    if user_id_str == "00000000-0000-0000-0000-000000000001":
        role = "student"
            
    if role == "faculty":
        # Allow viewing other faculty's assessments IF they are published OR if it's mine
        is_owner = str(assessment.created_by) == user_id_str
        if not is_owner and not assessment.is_published:
            raise HTTPException(status_code=403, detail="Not authorized to view this draft assessment")
    elif role == "student":
        student = db.query(Student).filter(Student.user_id == current_user["user_id"]).first()
        if not student and user_id_str == "00000000-0000-0000-0000-000000000001":
            # Manual mock student creation for emergency fallback
            from app.models.enums import BranchType, AcademicYear
            from app.models.auth_user import AuthUser
            mock_user = AuthUser(
                id=uuid.UUID(user_id_str),
                email="aarav@placement360.dev",
                full_name="Aarav Sharma",
                raw_app_meta_data={"role": "student"},
                email_confirmed_at=datetime.utcnow()
            )
            student = Student(
                user_id=uuid.UUID(user_id_str),
                roll_number="MOCK001",
                branch=BranchType.COMPUTER_SCIENCE,
                batch_year=2022,
                current_year=AcademicYear.FOURTH_YEAR
            )
            student.user = mock_user
            
        if student:
            branch_val = student.branch.value if hasattr(student.branch, 'value') else str(student.branch)
            year_val = student.current_year.value if hasattr(student.current_year, 'value') else str(student.current_year)
            
            # Reuse normalization logic for consistent access
            branch_map = {
                "computer_science": ["CS", "Computer Science", "CSE"],
                "information_technology": ["IT", "Information Technology"],
                "electronics": ["ECE", "Electronics", "ENTC"],
                "mechanical": ["Mech", "Mechanical"],
                "civil": ["Civil"],
                "electrical": ["EE", "Electrical"],
                "ai_ml": ["AI/ML", "AI", "ML"]
            }
            year_map = {
                "first": ["First Year", "1st Year"],
                "second": ["Second Year", "2nd Year"],
                "third": ["Third Year", "3rd Year"],
                "fourth": ["Fourth Year", "Final Year", "4th Year"]
            }
            
            student_branches = [branch_val]
            if branch_val in branch_map: student_branches.extend(branch_map[branch_val])
            
            student_years = [year_val, str(student.batch_year)]
            if year_val in year_map: student_years.extend(year_map[year_val])
            
            b_match = not assessment.target_branches
            if assessment.target_branches:
                for b in student_branches:
                    if b in assessment.target_branches:
                        b_match = True
                        break
            
            y_match = not assessment.target_years
            if assessment.target_years:
                for y in student_years:
                    if y in assessment.target_years:
                        y_match = True
                        break

            if not (b_match and y_match):
                raise HTTPException(status_code=403, detail="Assessment not assigned to you")
        else:
            raise HTTPException(status_code=403, detail="Student profile not found")
    else:
        raise HTTPException(status_code=403, detail="Not authorized")
            
    aqs = db.query(AssessmentQuestion).filter(
        AssessmentQuestion.assessment_id == assessment_id
    ).order_by(AssessmentQuestion.question_order).all()
    
    # Shuffle questions if enabled and it's a student viewing
    if role == "student" and assessment.shuffle_questions:
        import random
        # We use a seed based on user_id to keep the random order consistent for this student
        # so they don't get a new order on every refresh
        random_seed = int(str(current_user["user_id"]).replace("-", ""), 16) % (2**32)
        random.seed(random_seed)
        aqs_list = list(aqs)
        random.shuffle(aqs_list)
        aqs = aqs_list
    
    questions = []
    for aq in aqs:
        q = aq.question
        if q:
            # Build question detail object with all required fields
            is_coding = q.question_type == QuestionType.CODING
            is_mcq = q.question_type == QuestionType.MCQ
            
            options_data = q.options if is_mcq else None
            if options_data and role == "student":
                # Remove correct answers to prevent cheating
                scrubbed_options = []
                for opt in options_data:
                    safe_opt = dict(opt)
                    safe_opt.pop("isCorrect", None)
                    safe_opt.pop("is_correct", None)
                    scrubbed_options.append(safe_opt)
                options_data = scrubbed_options
                
            q_data = {
                "id": str(q.id),
                "question_order": aq.question_order,
                "title": q.title,
                "question_type": q.question_type.value if hasattr(q.question_type, 'value') else str(q.question_type),
                "difficulty": q.difficulty.value if hasattr(q.difficulty, 'value') else str(q.difficulty),
                "marks": aq.marks,
                "description": q.description,
                "input_format": q.input_format,
                "output_format": q.output_format,
                "constraints": q.constraints,
                "sample_test_cases": q.sample_test_cases if is_coding else [],
                "starter_code": q.starter_code if is_coding else None,
                "options": options_data,
                "time_limit_seconds": q.time_limit_seconds,
                "memory_limit_mb": q.memory_limit_mb
            }
            questions.append(q_data)
            
    # Use centralized formatter
    res_data = format_assessment_dict(assessment, db=db, current_user_id=user_id_str)
    
    # Add participants and scores for the detail view
    attempts = db.query(AssessmentAttempt).filter(AssessmentAttempt.assessment_id == assessment_id).all()
    submitted_attempts = [att for att in attempts if att.is_submitted]
    scores = [float(att.total_score) for att in submitted_attempts if att.total_score is not None]
    
    res_data["participants"] = len(set([str(att.student_id) for att in attempts]))
    
    raw_avg = (sum(scores) / len(scores)) if scores else 0.0
    res_data["avgScore"] = round(float(raw_avg / (assessment.total_marks if assessment.total_marks > 0 else 1.0) * 100.0), 1) if assessment.total_marks > 0 else 0.0
    res_data["highestScore"] = max(scores) if scores else 0.0
    res_data["lowestScore"] = min(scores) if scores else 0.0
    
    passed_attempts = [att for att in submitted_attempts if att.is_passed]
    res_data["passCount"] = len(passed_attempts)
    res_data["failCount"] = len(submitted_attempts) - len(passed_attempts)
    res_data["passRate"] = round(float(len(passed_attempts) / (len(scores) if scores else 1.0) * 100.0), 1) if scores else 0.0
        
    res_data["questions"] = questions
    
    # Analytics distribution using scores already calculated
    dist_buckets = {"0-20": 0, "21-40": 0, "41-60": 0, "61-80": 0, "81-100": 0}
    for s in scores:
        perc = (s / assessment.total_marks * 100.0) if assessment.total_marks > 0 else 0
        if perc <= 20:   dist_buckets["0-20"]   += 1
        elif perc <= 40: dist_buckets["21-40"]  += 1
        elif perc <= 60: dist_buckets["41-60"]  += 1
        elif perc <= 80: dist_buckets["61-80"]  += 1
        else:            dist_buckets["81-100"] += 1
    res_data["score_distribution"] = [{"range": k, "count": v} for k, v in dist_buckets.items()]

    # Topic breakdown as ARRAY (frontend expects {topic, scored, max, batchAvg} objects)
    # We build per-tag averages across all questions in the assessment
    topic_max: dict = {}
    topic_total_marks: dict = {}   # sum of marks available per tag
    for aq in aqs:
        q_tags = aq.question.tags if aq.question and aq.question.tags else []
        for tag in q_tags:
            topic_max[tag]          = topic_max.get(tag, 0) + (aq.marks or 0)
            topic_total_marks[tag]  = topic_max[tag]

    # Per-attempt scores grouped by tag
    tag_scores: dict = {}   # tag -> list of scores from each attempt
    for att in submitted_attempts:
        q_scores = att.question_scores or {}
        for aq2 in aqs:
            q2 = aq2.question
            if not q2: continue
            q_score = float(q_scores.get(str(q2.id), 0))
            for tag in (q2.tags or []):
                tag_scores.setdefault(tag, []).append(q_score)

    topic_breakdown_arr = []
    sorted_topics = sorted(topic_max.keys())
    for tag in sorted_topics:
        scores_for_tag = tag_scores.get(tag, [])
        avg_scored     = round(sum(scores_for_tag) / len(scores_for_tag), 1) if scores_for_tag else 0
        max_marks      = topic_max[tag]
        topic_breakdown_arr.append({
            "topic":    tag,
            "scored":   avg_scored,
            "max":      max_marks,
            "batchAvg": avg_scored,   # same as avg_scored for now; diverges once individual score is known
        })
    res_data["topic_breakdown"] = topic_breakdown_arr

    # Student results for faculty table view
    student_results = []
    sorted_by_score = sorted(submitted_attempts, key=lambda a: float(a.total_score or 0), reverse=True)
    for att in sorted_by_score:
        s_name = att.student.user.full_name if att.student and att.student.user else "Unknown"
        s_roll = att.student.roll_number if att.student else "N/A"
        student_results.append({
            "id":          str(att.id),
            "name":        s_name,
            "rollNumber":  s_roll,
            "score":       float(att.total_score) if att.total_score is not None else 0,
            "timeTaken":   att.time_taken_minutes,
            "status":      "pass" if att.is_passed else "fail",
            "tab_switches": att.tab_switch_count,
            "violation":    getattr(att, 'violation_flag', False),
            "violation_reason": getattr(att, 'violation_reason', ""),
        })
    res_data["student_results"] = student_results

    # ─── Student-specific personal result data ────────────────────────────────
    # Inject when caller is a student who has already attempted this assessment
    if role == "student" and student:
        my_attempt = db.query(AssessmentAttempt).filter(
            AssessmentAttempt.assessment_id == assessment_id,
            AssessmentAttempt.student_id    == student.id,
            AssessmentAttempt.is_submitted  == True            # noqa: E712
        ).order_by(AssessmentAttempt.submitted_at.desc()).first()

        if my_attempt:
            my_score_val = float(my_attempt.total_score) if my_attempt.total_score is not None else None

            # Rank calculation (among submitted attempts)
            rank = None
            if my_score_val is not None:
                rank = sum(1 for s in scores if s > my_score_val) + 1

            # Percentile
            percentile = None
            if scores and my_score_val is not None:
                below = sum(1 for s in scores if s < my_score_val)
                percentile = round((below / len(scores)) * 100, 1)

            # Batch average score
            batch_avg = round(sum(scores) / len(scores), 1) if scores else None

            # Per-topic personal breakdown (override batchAvg in topic_breakdown)
            my_q_scores = my_attempt.question_scores or {}
            my_tag_scores: dict = {}
            for aq2 in aqs:
                q2 = aq2.question
                if not q2: continue
                my_q_score = float(my_q_scores.get(str(q2.id), 0))
                for tag in (q2.tags or []):
                    my_tag_scores.setdefault(tag, [])
                    my_tag_scores[tag].append(my_q_score)

            personal_topic_breakdown = []
            for item in topic_breakdown_arr:
                tag        = item["topic"]
                my_scores_for_tag = my_tag_scores.get(tag, [])
                my_scored  = round(sum(my_scores_for_tag), 1) if my_scores_for_tag else 0
                personal_topic_breakdown.append({
                    "topic":    tag,
                    "scored":   my_scored,
                    "max":      item["max"],
                    "batchAvg": item["batchAvg"],
                })

            res_data["my_score"]       = my_score_val
            res_data["my_rank"]        = rank
            res_data["percentile"]     = percentile
            res_data["is_passed"]      = my_attempt.is_passed
            res_data["time_taken"]     = my_attempt.time_taken_minutes
            res_data["batch_avg_score"] = batch_avg
            res_data["topic_breakdown"] = personal_topic_breakdown if personal_topic_breakdown else topic_breakdown_arr

    res = {"success": True, "data": res_data}
    set_cached(cache_key, res, 60)
    return res


@router.patch("/{assessment_id}/publish", response_model=dict)
def publish_assessment(
    assessment_id: uuid.UUID,
    db: Session = Depends(get_db),
    current_faculty: Faculty = Depends(get_current_faculty)
):
    assessment = db.query(Assessment).filter(
        Assessment.id == assessment_id,
        Assessment.created_by == current_faculty.user_id
    ).first()
    
    if not assessment:
        raise HTTPException(status_code=404, detail="Assessment not found")
        
    if assessment.total_questions == 0:
        raise HTTPException(status_code=400, detail="Cannot publish assessment with no questions")
        
    assessment.is_published = True
    # Set status based on start_time
    now = datetime.utcnow()
    # Ensure now is aware if start_time is aware, though model use naive utc usually
    # But start_time here might be aware due to DateTime(timezone=True)
    if assessment.start_time:
        # Check if start_time is aware
        if assessment.start_time.tzinfo is not None:
             from datetime import timezone
             now = datetime.now(timezone.utc)
        
        if assessment.start_time > now:
            assessment.status = "scheduled"
        else:
            assessment.status = "active"
    else:
        assessment.status = "active"
    db.commit()
    # Aggressive cache invalidation
    invalidate_cache("assessment_")
    invalidate_cache("assessments_")
    invalidate_cache("student_dashboard:")
    invalidate_cache("faculty_dashboard:")
    
    return {"success": True, "message": "Assessment published successfully"}

@router.post("/{assessment_id}/questions", response_model=dict)
def add_question_to_assessment(
    assessment_id: uuid.UUID,
    data: AssessmentQuestionAdd,
    db: Session = Depends(get_db),
    current_faculty: Faculty = Depends(get_current_faculty)
):
    assessment = db.query(Assessment).filter(
        Assessment.id == assessment_id,
        Assessment.created_by == current_faculty.user_id
    ).first()
    
    if not assessment:
        raise HTTPException(status_code=404, detail="Assessment not found")
        
    question = db.query(Question).filter(Question.id == data.question_id).first()
    if not question:
        raise HTTPException(status_code=404, detail="Question not found")
        
    marks_to_add = data.marks if data.marks is not None else question.max_score
    order_to_use = data.question_order if data.question_order is not None else (assessment.total_questions + 1)
        
    aq = AssessmentQuestion(
        assessment_id=assessment.id,
        question_id=question.id,
        marks=marks_to_add,
        question_order=order_to_use
    )
    db.add(aq)
    
    assessment.total_questions += 1
    assessment.total_marks += marks_to_add
    
    db.commit()
    invalidate_cache("assessment_")
    invalidate_cache("assessments_")
    invalidate_cache("student_dashboard:")
    invalidate_cache("faculty_dashboard:")
    
    return {"success": True, "message": "Question added successfully"}

@router.get("/{assessment_id}/attempts", response_model=dict)
def get_assessment_attempts(
    assessment_id: uuid.UUID,
    db: Session = Depends(get_db),
    current_faculty: Faculty = Depends(get_current_faculty)
):
    """Get all student attempts for a specific assessment (Faculty only)."""
    assessment = db.query(Assessment).filter(
        Assessment.id == assessment_id,
        Assessment.created_by == current_faculty.user_id
    ).first()
    
    if not assessment:
        raise HTTPException(status_code=404, detail="Assessment not found")
        
    from sqlalchemy.orm import joinedload
    attempts = db.query(AssessmentAttempt).options(
        joinedload(AssessmentAttempt.student).joinedload(Student.user)
    ).filter(
        AssessmentAttempt.assessment_id == assessment_id
    ).all()
    
    data = []
    for att in attempts:
        # Calculate time taken in minutes
        time_taken = 0
        if att.started_at and att.submitted_at:
            delta = att.submitted_at - att.started_at
            time_taken = int(delta.total_seconds() / 60)
            
        score = float(att.total_score) if att.total_score is not None else 0.0
        passing_score = float(assessment.passing_marks or 0)
        max_score = float(att.max_score) if att.max_score is not None else float(assessment.total_marks or 0)
        
        percentage = 0.0
        if max_score > 0:
            percentage = round(float(score / max_score) * 100.0, 2)
            
        data.append({
            "id": str(att.id),
            "attempt_id": str(att.id),
            "student_id": str(att.student_id),
            "name": att.student.user.full_name if att.student and att.student.user else "Unknown student",
            "student_name": att.student.user.full_name if att.student and att.student.user else "Unknown student",
            "rollNumber": att.student.roll_number if att.student else "N/A",
            "roll_number": att.student.roll_number if att.student else "N/A",
            "score": score,
            "maxScore": max_score,
            "max_score": max_score,
            "percentage": percentage,
            "timeTaken": time_taken,
            "time_taken_minutes": time_taken,
            "status": "pass" if score >= passing_score else "fail",
            "is_passed": score >= passing_score,
            "tab_switch_count": att.tab_switch_count,
            "startedAt": att.started_at.isoformat() if att.started_at else None,
            "submittedAt": att.submitted_at.isoformat() if att.submitted_at else None,
            "submitted_at": att.submitted_at.isoformat() if att.submitted_at else None,
            "isSubmitted": att.is_submitted,
            "is_submitted": att.is_submitted,
            "attempt_number": att.attempt_number,
            "is_graded": att.is_graded
        })
        
    return {"success": True, "data": data}

@router.delete("/{assessment_id}/questions/{question_id}", response_model=dict)
def remove_question_from_assessment(
    assessment_id: uuid.UUID,
    question_id: uuid.UUID,
    db: Session = Depends(get_db),
    current_faculty: Faculty = Depends(get_current_faculty)
):
    assessment = db.query(Assessment).filter(
        Assessment.id == assessment_id,
        Assessment.created_by == current_faculty.user_id
    ).first()
    
    if not assessment:
        raise HTTPException(status_code=404, detail="Assessment not found")
        
    aq = db.query(AssessmentQuestion).filter(
        AssessmentQuestion.assessment_id == assessment.id,
        AssessmentQuestion.question_id == question_id
    ).first()
    
    if not aq:
        raise HTTPException(status_code=404, detail="Question mapping not found")
        
    assessment.total_questions -= 1
    assessment.total_marks -= aq.marks
    
    db.delete(aq)
    db.commit()
    
    invalidate_cache("assessment_")
    invalidate_cache("assessments_")
    invalidate_cache("student_dashboard:")
    invalidate_cache("faculty_dashboard:")
    
    return {"success": True, "message": "Question removed successfully"}

@router.post("/{assessment_id}/attempt/start", response_model=dict)
def start_attempt(
    assessment_id: uuid.UUID,
    db: Session = Depends(get_db),
    current_student: Student = Depends(get_current_student)
):
    assessment = db.query(Assessment).filter(Assessment.id == assessment_id).first()
    if not assessment:
        raise HTTPException(status_code=404, detail="Assessment not found")
        
    if not assessment.is_published:
        raise HTTPException(status_code=403, detail="Assessment is not published")

    # 1. Eligibility Check
    student_branch = normalize_branch(current_student.branch.value)
    assessment_branches = [normalize_branch(b) for b in (assessment.target_branches or [])]
    
    if assessment_branches and student_branch not in assessment_branches:
        raise HTTPException(status_code=403, detail="Your branch is not eligible for this assessment")
        
    student_year = normalize_year(current_student.current_year.value)
    assessment_years = [normalize_year(y) for y in (assessment.target_years or [])]
    
    if assessment_years and student_year not in assessment_years:
        raise HTTPException(status_code=403, detail="Your academic year is not eligible for this assessment")

    # 2. Max Attempts Check
    previous_attempts_count = db.query(AssessmentAttempt).filter(
        AssessmentAttempt.assessment_id == assessment_id,
        AssessmentAttempt.student_id == current_student.id
    ).count()
    
    active_attempt = db.query(AssessmentAttempt).filter(
        AssessmentAttempt.assessment_id == assessment_id,
        AssessmentAttempt.student_id == current_student.id,
        AssessmentAttempt.is_submitted == False
    ).first()

    if active_attempt:
        # Resume existing attempt
        attempt = active_attempt
    else:
        if previous_attempts_count >= assessment.max_attempts:
            raise HTTPException(status_code=400, detail="Maximum attempts reached")

        # 3. Time Window Check
        from datetime import timezone as tz
        now = datetime.now(tz.utc)
        now_naive = datetime.utcnow()
        
        def _to_aware(dt):
            """Ensure datetime is timezone-aware."""
            if dt is not None and dt.tzinfo is None:
                from datetime import timezone as tz2
                return dt.replace(tzinfo=tz2.utc)
            return dt
        
        start_cmp = _to_aware(assessment.start_time)
        end_cmp = _to_aware(assessment.end_time)
        
        if start_cmp and now < start_cmp:
            raise HTTPException(status_code=400, detail="Exam window not open")
        if end_cmp and now > end_cmp and not assessment.allow_late_submission:
            raise HTTPException(status_code=400, detail="Exam window has closed")

        # 4. Create Attempt
        attempt = AssessmentAttempt(
            assessment_id=assessment_id,
            student_id=current_student.id,
            attempt_number=previous_attempts_count + 1,
            started_at=datetime.utcnow(),
            is_submitted=False,
            is_graded=False,
            total_score=Decimal("0.0"),
            max_score=assessment.total_marks,
            tab_switch_count=0,
            questions_attempted=0,
            questions_correct=0
        )
        db.add(attempt)
        db.commit()
        db.refresh(attempt)
    
    # 5. Fetch questions
    aq_mappings = db.query(AssessmentQuestion).filter(
        AssessmentQuestion.assessment_id == assessment.id
    ).order_by(AssessmentQuestion.question_order).all()
    
    # Shuffle questions if enabled
    if assessment.shuffle_questions:
        import random
        random_seed = int(str(current_student.user_id).replace("-", ""), 16) % (2**32)
        random.seed(random_seed)
        mappings_list = list(aq_mappings)
        random.shuffle(mappings_list)
        aq_mappings = mappings_list
    
    questions_data = []
    for mapping in aq_mappings:
        q = mapping.question
        
        is_mcq = q.question_type == QuestionType.MCQ
        options_data = q.options if is_mcq else None
        
        # We MUST scrub correct answers before sending down to the client!
        if options_data:
            scrubbed_options = []
            for opt in options_data:
                safe_opt = dict(opt)
                safe_opt.pop("isCorrect", None)
                safe_opt.pop("is_correct", None)
                scrubbed_options.append(safe_opt)
            options_data = scrubbed_options
            
        questions_data.append({
            "id": str(q.id),
            "question_order": mapping.question_order,
            "title": q.title,
            "question_type": q.question_type.value if hasattr(q.question_type, 'value') else str(q.question_type),
            "difficulty": q.difficulty.value if hasattr(q.difficulty, 'value') else str(q.difficulty),
            "marks": mapping.marks,
            "description": q.description,
            "input_format": q.input_format,
            "output_format": q.output_format,
            "constraints": q.constraints,
            "sample_test_cases": q.sample_test_cases if q.question_type == QuestionType.CODING else [],
            "starter_code": q.starter_code if q.question_type == QuestionType.CODING else None,
            "options": options_data
        })
        
    return {
        "success": True,
        "data": {
            "attempt_id": str(attempt.id),
            "assessment_id": str(assessment.id),
            "attempt_number": attempt.attempt_number,
            "started_at": attempt.started_at.isoformat(),
            "time_limit_minutes": assessment.duration_minutes,
            "tab_switch_limit": assessment.tab_switch_limit,
            "questions": questions_data
        }
    }

@router.post("/{assessment_id}/attempt/submit", response_model=dict)
def submit_attempt(
    assessment_id: uuid.UUID,
    data: AttemptSubmitRequest,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_student: Student = Depends(get_current_student)
):
    attempt = db.query(AssessmentAttempt).filter(
        AssessmentAttempt.id == data.attempt_id,
        AssessmentAttempt.student_id == current_student.id
    ).first()

    if not attempt:
        raise HTTPException(status_code=404, detail="Attempt not found")
        
    if attempt.is_submitted:
        raise HTTPException(status_code=400, detail="Attempt already submitted")

    assessment = attempt.assessment
    
    # ── Security: Proctoring Check ───────────────────
    proctoring_violated = False
    if assessment.track_tab_switches:
        limit = assessment.tab_switch_limit or 3
        if data.tab_switch_count >= limit:
            logger.warning(f"submit_attempt: Proctoring violation for student {current_student.id}. Switches: {data.tab_switch_count}, Limit: {limit}")
            proctoring_violated = True
    
    total_score = Decimal("0.0")
    question_scores = {}
    
    for ans in data.answers:
        # Validate question_id is a proper UUID before querying.
        # The frontend may send integer IDs (e.g., "1") for mock questions.
        q_uuid = None
        q_order = None
        try:
            q_uuid = uuid.UUID(str(ans.question_id))
        except (ValueError, AttributeError):
            # Not a valid UUID — maybe it's an integer/order ID?
            try:
                q_order = int(str(ans.question_id))
                logger.info(f"submit_attempt: mapping integer ID {ans.question_id!r} to question_order={q_order}")
            except ValueError:
                logger.warning(f"submit_attempt: skipping invalid question_id={ans.question_id!r}")
                continue
        
        # Build query filter — if we have a UUID, search by it. Otherwise, search by assessment_id + question_order.
        if q_uuid:
            mapping = db.query(AssessmentQuestion).filter(
                AssessmentQuestion.assessment_id == assessment_id,
                AssessmentQuestion.question_id == q_uuid
            ).first()
        else:
            mapping = db.query(AssessmentQuestion).filter(
                AssessmentQuestion.assessment_id == assessment_id,
                AssessmentQuestion.question_order == q_order
            ).first()
        
        if not mapping:
            logger.warning(f"submit_attempt: could not find mapping for {'UUID ' + str(q_uuid) if q_uuid else 'Order ' + str(q_order)}")
            continue
            
        q = mapping.question
        score_awarded = Decimal("0.0")
        
        if q.question_type == QuestionType.MCQ:
            # Auto-grade MCQ
            # solution_code JSONB stores {"correct_option": 2}
            correct_option = None
            if q.solution_code:
                correct_option = q.solution_code.get("correct_option")
            
            if ans.selected_option is not None and ans.selected_option == correct_option:
                score_awarded = Decimal(str(mapping.marks))
                attempt.questions_correct += 1
            
            # SAVE SUBMISSION FOR MCQ TOO - so faculty can see the answer
            # NOTE: The submissions.language column is a PostgreSQL enum (programming_language)
            # that has no "mcq" value. We use PYTHON as a sentinel placeholder for MCQ rows.
            submission = Submission(
                student_id=current_student.id,
                question_id=q.id,
                assessment_attempt_id=attempt.id,
                language=ProgrammingLanguage.PYTHON,  # sentinel for MCQ — language field is NOT NULL
                source_code=str(ans.selected_option) if ans.selected_option is not None else "",
                status=SubmissionStatus.ACCEPTED,
                max_score=mapping.marks,
                score=score_awarded
            )
            db.add(submission)

            
            question_scores[str(q.id)] = float(score_awarded)
            total_score += score_awarded
        else:
            # Coding question - use Judge0 Background Task
            code_content = ans.source_code if ans.source_code else ""
            status = SubmissionStatus.PENDING
            
            submission = Submission(
                student_id=current_student.id,
                question_id=q.id,
                assessment_attempt_id=attempt.id,
                language=ans.language if ans.language else "python",
                source_code=code_content,
                status=status,
                max_score=mapping.marks,
                score=Decimal("0.0"),
                passed_test_cases=0,
                total_test_cases=len(q.hidden_test_cases or [])
            )
            db.add(submission)
            db.flush() # flush to get submission.id
            
            if code_content.strip():
                from app.integrations.judge0 import evaluate_coding_submission_bg
                background_tasks.add_task(
                    evaluate_coding_submission_bg,
                    str(submission.id),
                    code_content,
                    ans.language or "python",
                    q.hidden_test_cases or [],
                    float(mapping.marks)
                )
            
            # Default to 0 initially, Judge0 BG task will update attempt total_score
            question_scores[str(q.id)] = 0.0
            
    # ── Proctoring Finalization ──────────────────────
    if proctoring_violated:
        attempt.total_score = Decimal("0")
        attempt.is_passed = False
        attempt.violation_flag = True
        attempt.violation_reason = f"Exceeded tab switch limit of {limit}"
        # Still calculated for logs but effectively failed
        logger.warning(f"submit_attempt: Force failed attempt {attempt.id} due to violation.")
    else:
        attempt.is_passed = total_score >= (assessment.passing_marks or 0)
        attempt.total_score = total_score

    attempt.is_submitted = True
    attempt.submitted_at = datetime.utcnow()
    attempt.time_taken_minutes = data.time_taken_minutes
    attempt.tab_switch_count = data.tab_switch_count
    
    # Serialize proctoring events
    def _serialize_event(e):
        if isinstance(e, dict): return e
        try: return e.model_dump(mode='json')
        except: 
            try: return e.dict()
            except: return str(e)
    
    attempt.proctoring_events = [_serialize_event(e) for e in (data.proctoring_events or [])]
    attempt.questions_attempted = len(data.answers)
    attempt.question_scores = question_scores
    
    # Calculate percentage
    if attempt.max_score > 0:
        attempt.percentage = (attempt.total_score / Decimal(str(attempt.max_score))) * 100
    else:
        attempt.percentage = Decimal("0.0")
    
    db.commit()
    db.refresh(attempt)

    from app.services.analytics_service import calculate_readiness_score
    from app.models.student import Student
    student = db.query(Student).filter(Student.id == attempt.student_id).first()
    if student:
        calculate_readiness_score(student, db, save=True)

    # Invalidate student's assessment list cache
    branch_val = current_student.branch.value if hasattr(current_student.branch, 'value') else str(current_student.branch)
    year_val = current_student.current_year.value if hasattr(current_student.current_year, 'value') else str(current_student.current_year)
    cache_key = f"assessments_stu_{str(current_student.user_id)}_{branch_val}_{year_val}"
    invalidate_cache(cache_key)
    invalidate_cache(f"assessment_details_{str(assessment_id)}_{str(current_student.user_id)}")
    invalidate_cache("student_dashboard:")
    invalidate_cache("faculty_dashboard:")

    return {
        "success": True,
        "data": {
            "attempt_id": str(attempt.id),
            "total_score": float(attempt.total_score),
            "max_score": attempt.max_score,
            "percentage": float(attempt.percentage) if attempt.percentage else 0.0,
            "is_passed": attempt.is_passed,
            "questions_attempted": attempt.questions_attempted,
            "message": "Exam submitted successfully. Coding questions are being evaluated."
        }
    }


@router.get("/{assessment_id}/attempts/{attempt_id}", response_model=dict)
def get_assessment_attempt_detail(
    assessment_id: uuid.UUID,
    attempt_id: uuid.UUID,
    db: Session = Depends(get_db),
    current_faculty: Faculty = Depends(get_current_faculty)
):
    attempt = db.query(AssessmentAttempt).filter(
        AssessmentAttempt.id == attempt_id,
        AssessmentAttempt.assessment_id == assessment_id
    ).first()
    
    if not attempt:
        raise HTTPException(status_code=404, detail="Attempt not found")

    student = attempt.student
    student_name = student.user.full_name if student and student.user else "Unknown"

    # Get all questions in this assessment to show breakdown
    aq_mappings = db.query(AssessmentQuestion).filter(
        AssessmentQuestion.assessment_id == assessment_id
    ).order_by(AssessmentQuestion.question_order).all()

    # Get student submissions for this attempt
    submissions = {str(s.question_id): s for s in attempt.submissions}

    breakdown = []
    for mapping in aq_mappings:
        q = mapping.question
        score_awarded = attempt.question_scores.get(str(q.id), 0.0) if attempt.question_scores else 0.0
        
        student_answer = ""
        status = "not_attempted"
        
        sub = submissions.get(str(q.id))
        if sub:
            student_answer = sub.source_code
            status = sub.status.value
        elif q.question_type == QuestionType.MCQ:
            # We already started saving Submissions for MCQs in submit_attempt, 
            # so the sub block above will usually catch it.
            # But just in case, this is a fallback:
            if attempt.question_scores and str(q.id) in attempt.question_scores:
                status = "graded"
                
        is_correct = False
        if float(score_awarded) >= float(mapping.marks) and float(mapping.marks) > 0:
            is_correct = True
            
        if q.question_type == QuestionType.MCQ and status in ["graded", "accepted"]:
            status = "correct" if is_correct else "incorrect"
        
        breakdown.append({
            "question_id": str(q.id),
            "question_title": q.title,
            "question_type": q.question_type.value if hasattr(q.question_type, 'value') else str(q.question_type),
            "marks_awarded": float(score_awarded),
            "max_marks": mapping.marks,
            "student_answer": student_answer,
            "is_correct": is_correct,
            "status": status,
            "options": q.options if q.question_type == QuestionType.MCQ else None
        })

    return {
        "success": True,
        "data": {
            "attempt_id": str(attempt.id),
            "student_name": student_name,
            "roll_number": student.roll_number if student else "N/A",
            "total_score": float(attempt.total_score),
            "max_score": attempt.max_score,
            "percentage": float(attempt.percentage) if attempt.percentage else 0.0,
            "tab_switch_count": attempt.tab_switch_count,
            "violation": getattr(attempt, 'violation_flag', False),
            "violation_reason": getattr(attempt, 'violation_reason', ""),
            "proctoring_log": attempt.proctoring_events or [],
            "proctoring_events": attempt.proctoring_events or [],
            "time_taken_minutes": attempt.time_taken_minutes,
            "is_passed": attempt.is_passed,
            "is_submitted": attempt.is_submitted,
            "question_breakdown": breakdown
        }
    }


@router.get("/{assessment_id}/my-result", response_model=dict)
def get_my_result(
    assessment_id: uuid.UUID,
    db: Session = Depends(get_db),
    current_student: Student = Depends(get_current_student)
):
    """Get the student's own result for a specific assessment (Student only)."""
    attempt = db.query(AssessmentAttempt).filter(
        AssessmentAttempt.assessment_id == assessment_id,
        AssessmentAttempt.student_id == current_student.id,
        AssessmentAttempt.is_submitted == True
    ).order_by(AssessmentAttempt.submitted_at.desc()).first()
    
    if not attempt:
        return {"success": True, "data": None}
        
    student = attempt.student
    student_name = student.user.full_name if student and student.user else "Unknown"

    aq_mappings = db.query(AssessmentQuestion).filter(
        AssessmentQuestion.assessment_id == assessment_id
    ).order_by(AssessmentQuestion.question_order).all()

    submissions = {str(s.question_id): s for s in attempt.submissions}

    breakdown = []
    for mapping in aq_mappings:
        q = mapping.question
        score_awarded = attempt.question_scores.get(str(q.id), 0.0) if attempt.question_scores else 0.0
        
        student_answer = ""
        status = "not_attempted"
        
        sub = submissions.get(str(q.id))
        if sub:
            student_answer = sub.source_code
            status = sub.status.value
        elif q.question_type == QuestionType.MCQ:
            if attempt.question_scores and str(q.id) in attempt.question_scores:
                status = "graded"
                
        is_correct = False
        if float(score_awarded) >= float(mapping.marks) and float(mapping.marks) > 0:
            is_correct = True
            
        if q.question_type == QuestionType.MCQ and status in ["graded", "accepted"]:
            status = "correct" if is_correct else "incorrect"
        
        breakdown.append({
            "question_id": str(q.id),
            "question_title": q.title,
            "question_type": q.question_type.value if hasattr(q.question_type, 'value') else str(q.question_type),
            "marks_awarded": float(score_awarded),
            "max_marks": mapping.marks,
            "student_answer": student_answer,
            "is_correct": is_correct,
            "status": status,
            "options": q.options if q.question_type == QuestionType.MCQ else None
        })

    return {
        "success": True,
        "data": {
            "attempt_id": str(attempt.id),
            "student_name": student_name,
            "roll_number": student.roll_number if student else "N/A",
            "total_score": float(attempt.total_score),
            "max_score": attempt.max_score,
            "percentage": float(attempt.percentage) if attempt.percentage else 0.0,
            "tab_switch_count": attempt.tab_switch_count,
            "violation": getattr(attempt, 'violation_flag', False),
            "violation_reason": getattr(attempt, 'violation_reason', ""),
            "proctoring_log": attempt.proctoring_events or [],
            "proctoring_events": attempt.proctoring_events or [],
            "time_taken_minutes": attempt.time_taken_minutes,
            "is_passed": attempt.is_passed,
            "is_submitted": attempt.is_submitted,
            "question_breakdown": breakdown
        }
    }


@router.get("/{assessment_id}/export")
def export_assessment_results(
    assessment_id: uuid.UUID,
    format: str = Query("xlsx"),
    db: Session = Depends(get_db),
    current_faculty: Faculty = Depends(get_current_faculty),
):
    export_format = (format or "xlsx").strip().lower()
    if export_format not in {"xlsx", "csv"}:
        raise HTTPException(status_code=400, detail="format must be xlsx or csv")

    assessment = db.query(Assessment).filter(Assessment.id == assessment_id).first()
    if not assessment:
        raise HTTPException(status_code=404, detail="Assessment not found")

    if str(assessment.created_by) != str(current_faculty.user_id):
        raise HTTPException(status_code=403, detail="Not authorized to export this assessment")

    attempts = (
        db.query(AssessmentAttempt)
        .options(joinedload(AssessmentAttempt.student).joinedload(Student.user))
        .filter(
            AssessmentAttempt.assessment_id == assessment_id,
            AssessmentAttempt.is_submitted == True,  # noqa: E712
        )
        .order_by(AssessmentAttempt.total_score.desc(), AssessmentAttempt.submitted_at.asc())
        .all()
    )

    result_rows = [_build_result_row(rank=i + 1, attempt=attempt) for i, attempt in enumerate(attempts)]

    headers = [
        "Rank",
        "Student Name",
        "Roll No",
        "Branch",
        "Year",
        "CGPA",
        "Total Score",
        "Max Score",
        "Percentage",
        "Pass/Fail",
        "Time Taken (minutes)",
        "Tab Switches",
        "Submitted At",
    ]

    today_str = datetime.utcnow().strftime("%Y%m%d")
    safe_title = "-".join((assessment.title or "assessment").strip().split())

    if export_format == "csv":
        output = StringIO()
        writer = csv.DictWriter(output, fieldnames=headers)
        writer.writeheader()
        for row in result_rows:
            writer.writerow(row)

        csv_bytes = output.getvalue().encode("utf-8")
        filename = f"{safe_title}_results_{today_str}.csv"
        return StreamingResponse(
            iter([csv_bytes]),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Results"

    summary_sheet.append(headers)
    for col in range(1, len(headers) + 1):
        summary_sheet.cell(row=1, column=col).font = Font(bold=True)

    for row in result_rows:
        summary_sheet.append([row.get(header) for header in headers])

    for column_cells in summary_sheet.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        summary_sheet.column_dimensions[column].width = min(max_length + 2, 50)

    has_breakdown = any((attempt.question_scores or {}) for attempt in attempts)
    if has_breakdown:
        breakdown_sheet = workbook.create_sheet("Question Breakdown")
        breakdown_headers = ["Attempt ID", "Student Name", "Roll No", "Question ID", "Score"]
        breakdown_sheet.append(breakdown_headers)
        for col in range(1, len(breakdown_headers) + 1):
            breakdown_sheet.cell(row=1, column=col).font = Font(bold=True)

        for attempt in attempts:
            question_scores = attempt.question_scores or {}
            student_name = attempt.student.user.full_name if attempt.student and attempt.student.user else "Unknown"
            roll_no = attempt.student.roll_number if attempt.student else "N/A"
            for question_id, score in question_scores.items():
                breakdown_sheet.append(
                    [
                        str(attempt.id),
                        student_name,
                        roll_no,
                        str(question_id),
                        float(score) if score is not None else 0.0,
                    ]
                )

        for column_cells in breakdown_sheet.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            breakdown_sheet.column_dimensions[column].width = min(max_length + 2, 60)

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    filename = f"{safe_title}_results_{today_str}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.delete("/{assessment_id}", response_model=dict)
def delete_assessment(
    assessment_id: uuid.UUID,
    db: Session = Depends(get_db),
    current_faculty: Faculty = Depends(get_current_faculty)
):
    """Delete an assessment and cascade delete all its dependencies (Faculty only)."""
    assessment = db.query(Assessment).filter(Assessment.id == assessment_id).first()
    
    if not assessment:
        raise HTTPException(status_code=404, detail="Assessment not found")
        
    if str(assessment.created_by) != str(current_faculty.user_id):
        raise HTTPException(status_code=403, detail="Not authorized to delete this assessment")
        
    # Delete cascade exactly as requested
    
    # 1. Delete all submissions for attempts belonging to this assessment
    attempts = db.query(AssessmentAttempt.id).filter(AssessmentAttempt.assessment_id == assessment_id).all()
    attempt_ids = [a.id for a in attempts]
    
    if attempt_ids:
        db.query(Submission).filter(Submission.assessment_attempt_id.in_(attempt_ids)).delete(synchronize_session=False)
        
    # 2. Delete all rows in assessment_attempts where assessment_id = id
    db.query(AssessmentAttempt).filter(AssessmentAttempt.assessment_id == assessment_id).delete(synchronize_session=False)
    
    # 3. Delete all rows in assessment_questions where assessment_id = id
    db.query(AssessmentQuestion).filter(AssessmentQuestion.assessment_id == assessment_id).delete(synchronize_session=False)
    
    # 4. Delete the assessment itself
    db.delete(assessment)
    db.commit()
    
    # Clean cache
    invalidate_cache("assessment_")
    invalidate_cache("assessments_")
    invalidate_cache("student_dashboard:")
    invalidate_cache("faculty_dashboard:")
    
    return {"success": True, "message": "Assessment deleted successfully"}
